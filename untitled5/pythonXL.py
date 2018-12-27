#파일을 읽어들이는 내장함수인 load_workbook을 불러옵니다.
from openpyxl import load_workbook
from openpyxl import Workbook

#파일이름이 angel.xlsx인 파일을 불러옵니다.
angelEx=load_workbook(filename='angel.xlsx')
#불러온 엑셀 파일 중 데이터를 찾을 sheet의 이름을 입력합니다.
sheet1 = angelEx['Sheet1']
#활성화되어 있는 시트를 찾습니다.
sheet2 = angelEx.active
#Sheet1의 D4의 값을 출력합니다.
print(sheet1['D4'].value)
print(sheet2['D4'].value)

#특정 범위의 데이터(번호와 성인명)만 불러오겠습니다.
angels = []
#루프문을 이용해 sheet의 여러 행에 있는 데이터를 불러옵니다.
for i in sheet1.rows:
    번호 = i[0].value
    성인명 = i[2].value
    angel = (번호, 성인명)
    angels.append(angel)

# print(angels)

wb = Workbook()
sheettemp = wb.create_sheet("temp")

k = 1
for i in sheet1.rows:
    if k % 2 == 0:
        for x in range(1, 6):
            sheettemp.cell(k, x).value = i[x].value
    k = k + 1

# 이건 리스트로 만들어서 하나씩 출력이 된다.
for row in sheet1.rows:
    print([col.value for col in row])

# 이건 한줄씩 출력이 됨
# for row in sheet1.rows:
#     for col in row:
#         print(col.value)

filename = 'even.xlsx'
wb.save(filename)

